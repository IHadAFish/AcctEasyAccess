from flask import Flask, request, jsonify, send_file
from celery import Celery
from celery.schedules import crontab
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os, redis, json, tempfile, glob

from database_manager import DatabaseManager

load_dotenv()

app = Flask(__name__, static_folder='static')
app.config['CELERY_BROKER_URL'] = f"redis://{os.getenv('REDIS_HOST')}:{os.getenv('REDIS_PORT')}/{os.getenv('REDIS_CELERY_DB')}"

celery = Celery(app.name, broker=app.config['CELERY_BROKER_URL'])
celery.conf.update(
    app.config,
    beat_schedule={
        'refresh-reminders-hourly': {
            'task': 'app.routine_tasks',
            'schedule': crontab(minute=0, hour='*'), # Run at the start of every hour
        },
    }
)

redis_client = redis.Redis(host=os.getenv('REDIS_HOST'), port=int(os.getenv('REDIS_PORT')), db=int(os.getenv('REDIS_CACHE_DB'))) # Fixed typo: getevn -> getenv

#####################
#
# Helper functions
#
#####################
def get_cached_plan(plan_id):
    # Use ':' separator consistent with cache setting
    cached_plan = redis_client.get(f'plan:{plan_id}')
    if cached_plan:
        return json.loads(cached_plan.decode('utf-8'))
    return None

def load_business_cache():
    db_manager = DatabaseManager()

    business_data = db_manager.get_businesses()
    for name, id in business_data.items():
        redis_client.set(f'business_{name}', id)
        redis_client.set(f'business:id:{id}', name)

def warm_cache():
    load_business_cache()

    # Reminder cache setup
    # rem:detail:{reminder_id} - details for that reminder
    # rem:global:{market}:{business_id} -
    #       market - could be "USCN" or "CACN", if a reminder have both values, it should be present in both
    #       business_id - the value is "all" if is_all_business is true
    #       the value for this would be a set of reminder_ids
    # rem:{reminder_id}:{acct_id} - hash with two fields, only for unexpired event:
    #       is_processed: current status for this reminder of this user
    #       last_processed_at: last time the process status changed, note that the last_processed_at in acct_to_reminders table would only be changed when is_processed is set to true
    # rem:acct_list - a set of acct ids which have at least one active non global reminder
    # rem:acct:{acct_id} - a set of non global reminders for a specific user
    #
    # Cache/invalidate strategy:
    # 1. Periodically(hourly) retrieve unexpired reminders from db(only the ones modified after last retrieval?)
    # 2. Insert/overwrite the new reminders into rem:global and rem:detail, outdated reminders are not removed
    # 3. When an remider is being served, check these and remove from rem:global, rem:detail and rem:{reminder_id}:{acct_id}
    #       - Check whether global category is correct(incase of update) 
    #       - Check whether its outdated
    # 4. When an acct's reminder list is being retrieved, if the result have no non global reminder, remove the user from rem:acct_list
    # 5. When an reminder is being added/updated:
    #       - Add/overwrite the reminder detail in rem:detail
    #       - If the reminder is global add/overwrite the reminder_id in rem:global(Potential prior reminder under other category is ignored)
    #       - If the reminder is non global, add all acct_id to rem:acct_list(add reminder_id to all existing rem:acct:{acct_id})
    #               And flush rem:{reminder_id}:{acct_id}
    # 6. When an action is done to a reminder, insert/overwrite to rem:{reminder_id}:{acct_id}
    # 
    # Cache warming:
    # 1. Load all unexpired reminders to rem:detail
    # 2. Among them, load global reminder_id to rem:global:{market}:{business_id}
    # 3. Load all acct_reminders for unexpired reminders to rem:{reminder_id}:{acct_id}
    # 4. Load all acct with non global reminders to rem:acct_list
    # 5. No action is done to rem:acct:{acct_id}, it's load lazily
    # 
    # A request for remider list of an acct should require: acct_id, market, business
    # Any action performed on a reminder should consist of an update to server
    # 
    # Alteration: CN_Num
    # 48k CN num are loaded into spe:cn_nums set, updated on beat 

    refresh_plans.delay(initial=True)
    refresh_reminders.delay(initial=True)
    load_CN_num.delay(initial=True)

def _validate_and_prepare_reminder_data(reminder_data, acct_ids=None):
    """
    Validates reminder data, sets defaults, and prepares it for insertion.
    Raises ValueError if validation fails.
    Returns the processed reminder dictionary.
    """
    # Rule: title and content must not be None
    if reminder_data.get('title') is None or reminder_data.get('content') is None:
        raise ValueError("Reminder title and content cannot be None.")

    # Rule: market validation and default
    market = reminder_data.get('market')
    valid_markets = {'USCN', 'CACN'}
    processed_market_list = [] # Store the validated list here
    if market is None or not market:
        processed_market_list = ['USCN', 'CACN']
    elif isinstance(market, str): # Handle comma-separated string input if necessary
        processed_market_list = [m.strip() for m in market.split(',') if m.strip() in valid_markets]
        if not processed_market_list:
             raise ValueError(f"Invalid market value provided: {market}. Must be 'USCN' or 'CACN'.")
    elif isinstance(market, list):
        processed_market_list = [m for m in market if m in valid_markets]
        if not processed_market_list: # Handle case of empty list after filtering
             processed_market_list = ['USCN', 'CACN']
        # Ensure no duplicates
        processed_market_list = sorted(list(set(processed_market_list)))
    else:
        raise ValueError("Market must be a list or comma-separated string.")
    reminder_data['market'] = processed_market_list # Update with the clean list

    # Rule: start_date default
    if not reminder_data.get('start_date'):
        # Use datetime object for consistency, DB layer handles conversion
        reminder_data['start_date'] = datetime.fromtimestamp(0)
    # Ensure it's a datetime object if provided
    elif isinstance(reminder_data['start_date'], str):
         try:
            reminder_data['start_date'] = datetime.fromisoformat(reminder_data['start_date'].replace('Z', '+00:00')) # Handle Z timezone
         except ValueError:
             raise ValueError("Invalid start_date format. Use ISO format (YYYY-MM-DDTHH:MM:SS).")

    # Rule: expire_date default
    if not reminder_data.get('expire_date'):
        # Use datetime object
        reminder_data['expire_date'] = datetime(5000, 1, 1)
    # Ensure it's a datetime object if provided
    elif isinstance(reminder_data['expire_date'], str):
         try:
            reminder_data['expire_date'] = datetime.fromisoformat(reminder_data['expire_date'].replace('Z', '+00:00')) # Handle Z timezone
         except ValueError:
             raise ValueError("Invalid expire_date format. Use ISO format (YYYY-MM-DDTHH:MM:SS).")
    
    if not reminder_data.get('recurrent_interval', False):
        reminder_data['recurrent_interval'] = 0

    # Rule: is_global checks
    is_global = reminder_data.get('is_global')
    if is_global is None:
        raise ValueError("is_global must be provided (True or False).")
    # Ensure boolean
    is_global = bool(is_global)
    reminder_data['is_global'] = is_global


    if is_global:
        # Rule: is_all_business must have a value if is_global is true
        is_all_business = reminder_data.get('is_all_business')
        if is_all_business is None:
            raise ValueError("is_all_business must be provided (True or False) when is_global is True.")
        # Ensure boolean
        is_all_business = bool(is_all_business)
        reminder_data['is_all_business'] = is_all_business

        # Rule: business_name validation if is_all_business is false
        if not is_all_business:
            business_names = reminder_data.get('business_name') # Expecting 'business_name' from input
            if not business_names or not isinstance(business_names, list):
                raise ValueError("business_name list must be provided and non-empty when is_global is True and is_all_business is False.")

            business_ids = []
            missing_businesses = []
            for name in business_names:
                cached_id = redis_client.get(f'business_{name}')
                if cached_id:
                    business_ids.append(int(cached_id.decode('utf-8')))
                else:
                    missing_businesses.append(name)

            # Optional: Attempt cache refresh if misses occur
            if missing_businesses:
                 print(f"Cache miss for businesses: {missing_businesses}. Attempting refresh.")
                 load_business_cache() # Ensure this function is accessible
                 for name in list(missing_businesses): # Iterate over a copy
                     cached_id = redis_client.get(f'business_{name}')
                     if cached_id:
                         business_ids.append(int(cached_id.decode('utf-8')))
                         missing_businesses.remove(name)

            if missing_businesses:
                raise ValueError(f"Could not find business IDs for: {', '.join(missing_businesses)}")

            reminder_data['business_ids'] = business_ids # Add the translated IDs
            reminder_data.pop('business_name', None) # Remove original names

        # Ensure acct is not present for global reminders
        reminder_data.pop('acct', None)

    else: # is_global is False
        # Rule: acct list must not be empty
        if not acct_ids: # Check the passed acct_ids argument
            raise ValueError("Account list (acct_ids) must be provided when is_global is False.")
        reminder_data['acct'] = acct_ids # Add acct list for DB layer

        # Rule: is_all_business should be None
        # Set to None for clarity before DB insertion, schema allows NULL
        reminder_data['is_all_business'] = None

        reminder_data.pop('business_ids', None)
        reminder_data.pop('business_name', None)

    return reminder_data

#####################
#
# Celery beats
#
#####################
@celery.task(bind=True)
def routine_tasks(self):
    refresh_reminders.delay()
    load_CN_num.delay()
    refresh_plans.delay()

@celery.task(bind=True)
def load_CN_num(self, initial=None):
    try:
        lookback_minutes = 65
        if initial:
            since_time = datetime.fromtimestamp(0)
        else:
            since_time = datetime.now() - timedelta(minutes=lookback_minutes)

        if datetime.fromtimestamp(os.path.getmtime(os.getenv("CN_NUM_PATH"))) < since_time:
            return "CN num not loaded, no changes since last refresh."

        wb = load_workbook(os.getenv("CN_NUM_PATH"))
        ws = wb.active

        cn_nums = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0]:
                cn_nums.append(int(str(row[0])))
        
        redis_client.sadd("spe:cn_nums", *cn_nums)

        return f"CN num loaded, {len(cn_nums)} total."
    except Exception as e:
        raise Exception("Load CN nums failed")

@celery.task(bind=True)
def refresh_plans(self, initial=False):
    try:
        db_manager = DatabaseManager()

        if initial:
            last_modified = datetime.fromtimestamp(0)
        else:
            last_modified = datetime.now() - timedelta(minutes=65)

        plan_ids = db_manager.get_all_plan_ids(last_modified=last_modified)
        if not plan_ids:
            return f"Plan refreshed, 0 total"

        plans = db_manager.get_plan(plan_ids)
        for plan in plans:
            plan_dict = {
                'plan_id': plan[0],
                'name': plan[1],
                'contract': plan[2],
                'business': plan[3],
                'fee': plan[4],
                'discount': plan[5],
                'note': plan[6]
            }
            redis_client.set(f'plan:{plan[0]}', json.dumps(plan_dict))

        return f"Plan refreshed, {len(plans)} total"
    except Exception as e:
        raise Exception(f"Error during cache warming: {e}")

@celery.task(bind=True)
def refresh_reminders(self, initial=False):
    """
    Refreshes the reminder cache by fetching recent updates from the database.
    This task is intended to be run periodically by Celery Beat.
    """
    db_manager = DatabaseManager()
    lookback_minutes = 65

    if initial:
        acct_reminders = db_manager.get_acct_reminder_status()
        since_time = datetime.fromtimestamp(0)
    else:
        since_time = datetime.now() - timedelta(minutes=lookback_minutes)
        acct_reminders = db_manager.get_acct_reminder_status(last_modified=since_time)

    for item in acct_reminders:
        if not item["last_processed_at"]:
            item["last_processed_at"] = datetime.fromtimestamp(0)

        redis_client.hset(f"rem:{item['reminder_id']}:{item['acct_id']}",
            mapping={"is_processed": item["is_processed"], "last_processed_at": item["last_processed_at"].isoformat()}
        )
        redis_client.sadd("rem:acct_list", item["acct_id"])

    # Fetch and update reminder details and global sets
    global_reminders_to_map = []
    reminders_to_update = db_manager.get_unexpired_reminders(since_time) # Fetch recently modified/unexpired

    pipeline = redis_client.pipeline()

    # Process reminders: Convert market string to list and format dates
    processed_reminders = []
    for rem in reminders_to_update:
        # Convert market string to list
        if isinstance(rem.get("market"), str):
            rem["market"] = [mkt for mkt in rem["market"].split(',') if mkt] # Split and filter empty strings
        elif rem.get("market") is None:
             rem["market"] = []

        rem["start_date"] = rem["start_date"].isoformat()
        rem["expire_date"] = rem["expire_date"].isoformat()

        processed_reminders.append(rem)

    # Separate logic for global reminders using the processed list
    global_reminders_to_map = []
    for rem in processed_reminders:
        rem_id = rem['id']
        if rem["is_global"]:
            if rem["is_all_business"]:
                for market in rem["market"]:
                    pipeline.sadd(f"rem:global:{market}:all", rem_id)
            else:
                global_reminders_to_map.append(rem)

    # Fetch business IDs for the relevant global reminders
    if global_reminders_to_map:
        rem_to_businesss = db_manager.get_business_ids_for_reminders([item["id"] for item in global_reminders_to_map])
        for rem in global_reminders_to_map:
            rem_id = rem["id"]
            business_ids = rem_to_businesss.get(rem_id, [])
            for b_id in business_ids:
                for m in rem["market"]:
                    pipeline.sadd(f"rem:global:{m}:{b_id}", rem_id)

    for rem in reminders_to_update:
        # Serialize market list to JSON before caching
        if isinstance(rem.get('market'), list):
            rem['market'] = json.dumps(rem['market'])
        for k in ['is_global', 'is_all_business']:
            if isinstance(rem.get(k), bool) or not rem.get(k):
               rem[k] = int(bool(rem.get(k, 0)))
        
        pipeline.hset(f"rem:detail:{rem['id']}", mapping=rem)

    pipeline.execute()
    return f"Refreshed reminder cache at {datetime.now()}, {len(reminders_to_update)} total, {len(acct_reminders)} items"

#####################
#
# Celery tasks
#
#####################
@celery.task(bind=True)
def sink_plan_to_db(self, plans):
    failed_count = 0
    failed_reason = ""
    processed_plan = []

    try:
        for plan in plans:
            try:
                for f in ["plan_id", "name", "business_name", "contract"]:
                    if f not in plan.keys():
                        raise Exception(f'{plan} missing field: {f}')

                for f in ["fee", "discount", "note"]:
                    if f not in plan.keys():
                        plan[f] = ""

                business_name = plan["business_name"]
                business_id = redis_client.get(f'business_{business_name}')
                if business_id:
                    plan['business_id'] = int(business_id.decode('utf-8'))
                    plan.pop('business_name', None)  # Remove business_name
                else:
                    # Cache miss: refresh the business cache
                    load_business_cache()
                    business_id = redis_client.get(f'business_{business_name}')
                    if business_id:
                        plan['business_id'] = int(business_id.decode('utf-8'))
                        plan.pop('business_name', None)  # Remove business_name
                    else:
                        raise Exception(f'未找到业务{business_name}')
                    
                processed_plan.append(plan)
            except Exception as e:
                failed_count += 1
                failed_reason += e
                continue

        db_manager = DatabaseManager()

        db_manager.update_plan(processed_plan)

        for plan in processed_plan:
            if redis_client.exists(f"plan:{plan['plan_id']}"):
                redis_client.set(f"plan:{plan['plan_id']}", json.dumps(plan))

        return f"{len(processed_plan)} processed, {failed_count} failed. Reason:\n{failed_reason}"
    except Exception as e:
        self.update_state(state="FAILURE")
        raise Exception(f'Error processing db operation: {str(e)}')
    
@celery.task(bind=True)
def get_plan_info(self, plan_id):
    try:
         
        cached_plan = get_cached_plan(plan_id)
        if cached_plan:
            return cached_plan

        db_manager = DatabaseManager()

        plan = db_manager.get_plan([plan_id])[0]
        
        plan_dict = {
            'plan_id': plan[0],
            'name': plan[1],
            'contract': plan[2],
            'business': plan[3],
            'fee': plan[4],
            'discount': plan[5],
            'note': plan[6]
        }
        redis_client.set(f'plan:{plan_id}', json.dumps(plan_dict))

        return plan_dict
        
    except Exception as e:
        self.update_state(state="FAILURE")
        raise Exception(f'Unable to load plan: {plan_id}')

@celery.task(bind=True)
def process_excel_data(self, file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = ["plan_id", "name", "contract", "business_name", "fee", "discount", "note"]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(header, row)))

        #sink_plan_to_db.delay(data)
        return data
    except Exception as e:
        self.update_state(state="FAILURE")
        raise Exception(f'Error processing plan: {str(e)}')
    finally:
        os.remove(file_path)

@celery.task(bind=True)
def get_acct_from_form(self, file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        acct_ids = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] and isinstance(row[0], int):
                acct_ids.append(row[0])

        return acct_ids
    except Exception as e:
        self.update_state(state="FAILURE")
        raise Exception(f'Error getting acct_id: {str(e)}')
    finally:
        os.remove(file_path)

@celery.task(bind=True)
def insert_reminder_task(self, acct_ids, reminder_data):
    db_manager = DatabaseManager()
    try:
        # Validate and prepare data using the helper function
        # Pass acct_ids only if the reminder might be non-global
        # A bit redundant check, but ensures acct_ids are passed correctly based on intent
        passed_acct_ids = acct_ids if not reminder_data.get('is_global', False) else None
        validated_reminder = _validate_and_prepare_reminder_data(reminder_data.copy(), acct_ids=passed_acct_ids) # Use copy to avoid modifying original dict if task retries

        # Insert into database
        reminder_id = db_manager.insert_reminder(validated_reminder)
        if reminder_id is False:
            # The insert_reminder method should ideally raise specific errors
            self.update_state(state="FAILURE")
            raise Exception(f"Database failed to insert reminder: {validated_reminder.get('title')}")

        validated_reminder["id"] = reminder_id # Add the returned ID

        # --- Cache Update Logic ---
        pipeline = redis_client.pipeline()

        # Clear potentially outdated status entries for this reminder
        for key in redis_client.scan_iter(f"rem:{reminder_id}:*"):
            pipeline.delete(key)

        # Update cache based on reminder type (Global vs Account-specific)
        if validated_reminder["is_global"]:
            if validated_reminder["is_all_business"]:
                for m in validated_reminder["market"]:
                    pipeline.sadd(f"rem:global:{m}:all", reminder_id)
            else:
                for m in validated_reminder["market"]:
                    for b_id in validated_reminder.get("business_ids", []):
                        pipeline.sadd(f"rem:global:{m}:{b_id}", reminder_id)
        else:
            if acct_ids:
                pipeline.sadd("rem:acct_list", *acct_ids)
                # Add reminder_id to existing per-account sets
                for acct in acct_ids:
                    if redis_client.exists(f"rem:acct:{acct}"):
                        pipeline.sadd(f"rem:acct:{acct}", reminder_id)
                # Fetch and cache initial status for the new associations
                for a in acct_ids:
                     pipeline.hset(f"rem:{reminder_id}:{a}",
                         mapping={
                             "is_processed": 0, 
                             "last_processed_at": 0
                         }
                     )

        # Cache the reminder detail
        # Prepare for JSON: Convert datetime objects and market list
        cacheable_reminder = validated_reminder.copy()
        if isinstance(cacheable_reminder.get('market'), list):
            cacheable_reminder['market'] = json.dumps(cacheable_reminder['market'])
        if isinstance(cacheable_reminder.get('start_date'), datetime):
            cacheable_reminder['start_date'] = cacheable_reminder['start_date'].isoformat()
        if isinstance(cacheable_reminder.get('expire_date'), datetime):
            cacheable_reminder['expire_date'] = cacheable_reminder['expire_date'].isoformat()
        for k in ['is_global', 'is_all_business']:
            cacheable_reminder[k] = int(bool(cacheable_reminder.get(k, 0)))
        # Remove keys not needed/suitable for direct caching if necessary (e.g., 'acct')
        cacheable_reminder.pop('acct', None)
        cacheable_reminder.pop('business_ids', None)
        cacheable_reminder.pop('business_names', None)
        
        pipeline.hset(f"rem:detail:{reminder_id}", mapping=cacheable_reminder)

        pipeline.execute()

        return f"Insertion successful: {reminder_id}"
    except ValueError as ve: # Catch validation errors
        self.update_state(state="FAILURE")
        # Re-raise or handle specific validation errors
        raise Exception(f'Invalid reminder data: {str(ve)}')
    except Exception as e:
        self.update_state(state="FAILURE")
        # Log the full error for debugging
        print(f"Error in insert_reminder_task: {type(e).__name__} - {str(e)}")
        raise Exception(f'Error processing reminder insertion: {str(e)}')

@celery.task(bind=True)
def get_reminders_for_acct(self, acct_id, market, business, cn_num=None):
    db_manager = DatabaseManager()
    try:
        reminder_list = []

        cached_id = redis_client.get(f'business_{business}')
        if cached_id:
            business_id = int(cached_id.decode('utf-8'))
        
        # --- Global reminders ---
        for k in [f"rem:global:{market}:all", f"rem:global:{market}:{business_id}"]:
            if redis_client.exists(k):
                reminder_ids = [rem_id.decode('utf-8') for rem_id in redis_client.smembers(k)]
                pipeline = redis_client.pipeline()
                for rem_id in reminder_ids:
                    pipeline.hgetall(f"rem:detail:{rem_id}")
                results = pipeline.execute()

                cleanup_pipeline = redis_client.pipeline()
                for i, rem_data_bytes in enumerate(results):
                    # Decode the hash data fetched by the pipeline
                    rem_data = {key.decode('utf-8'): value.decode('utf-8') if isinstance(value, bytes) else value
                                for key, value in rem_data_bytes.items()}

                    if not rem_data:
                        cleanup_pipeline.srem(k, reminder_ids[i])
                        continue

                    # Deserialize market JSON string from cache
                    market_json = rem_data.get('market')
                    market_list = []
                    if isinstance(market_json, str):
                        try:
                            market_list = json.loads(market_json)
                        except json.JSONDecodeError:
                            market_list = []
                    rem_data['market'] = market_list

                    if datetime.fromisoformat(rem_data["expire_date"]) < datetime.now():
                        cleanup_pipeline.delete(f"rem:detail:{reminder_ids[i]}")
                        cleanup_pipeline.srem(k, reminder_ids[i])
                    elif market not in market_list:
                         cleanup_pipeline.srem(k, reminder_ids[i])
                    else:
                        reminder_list.append(rem_data)
                cleanup_pipeline.execute()

        # --- Account-specific reminders ---
        if redis_client.sismember("rem:acct_list", acct_id):
            if redis_client.exists(f"rem:acct:{acct_id}"):
                reminder_ids = [rem_id.decode('utf-8') for rem_id in redis_client.smembers(f"rem:acct:{acct_id}")]

                pipeline = redis_client.pipeline()
                for rem_id in reminder_ids:
                    pipeline.hgetall(f"rem:detail:{rem_id}")
                results = pipeline.execute()

                cleanup_pipeline = redis_client.pipeline()
                for i, rem_data_bytes in enumerate(results):
                    # Decode bytes to strings
                    rem_data = {key.decode('utf-8'): value.decode('utf-8') if isinstance(value, bytes) else value for key, value in rem_data_bytes.items()}

                    if not rem_data:
                        cleanup_pipeline.srem(f"rem:acct:{acct_id}", reminder_ids[i])
                        continue

                    # Deserialize market JSON string from cache (same logic as above)
                    market_json = rem_data.get('market')
                    market_list = []
                    if isinstance(market_json, str):
                        try:
                            market_list = json.loads(market_json)
                            if not isinstance(market_list, list): market_list = []
                        except json.JSONDecodeError:
                            market_list = []
                    rem_data['market'] = market_list

                    # Perform checks
                    if datetime.fromisoformat(rem_data["expire_date"]) < datetime.now():
                        cleanup_pipeline.delete(f"rem:detail:{rem_data['id']}")
                        cleanup_pipeline.srem(f"rem:acct:{acct_id}", rem_data['id'])
                    else:
                        reminder_list.append(rem_data) # Append data with market as list
                cleanup_pipeline.execute()
            else:
                # Fallback to database if rem:acct:{acct_id} doesn't exist
                temp_rems = db_manager.get_reminders_by_acct(acct_id)
                if not temp_rems:
                    redis_client.srem("rem:acct_list", acct_id)
                else:
                    decoded_rems = []
                    for rem in temp_rems:
                        decoded_rem = {}
                        for key, value in rem.items():
                            if key == 'market':
                                # Convert comma-separated string from DB to list
                                if isinstance(value, str):
                                    decoded_rem[key] = [mkt for mkt in value.split(',') if mkt]
                                else:
                                    decoded_rem[key] = [] # Handle None or unexpected types
                            elif isinstance(value, datetime):
                                decoded_rem[key] = value.isoformat()
                            else:
                                decoded_rem[key] = value
                        decoded_rems.append(decoded_rem)
                    redis_client.sadd(f"rem:acct:{acct_id}", *[r["id"] for r in decoded_rems])
                    redis_client.setex(f"rem:acct:{acct_id}", 1200)
                    reminder_list.extend(decoded_rems) # Add processed reminders

        # Add status from rem:{rem_id}:{acct_id} hash
        temp_pipe = redis_client.pipeline()
        for rem in reminder_list: # Now rem['market'] should consistently be a list
            temp_pipe.hgetall(f"rem:{rem['id']}:{acct_id}")
        results = temp_pipe.execute()
        results = [{key.decode('utf-8'): value.decode('utf-8') if isinstance(value, bytes) else value for key, value in r.items()} for r in results] 

        for rem, stat in zip(reminder_list, results):
            # Assign the status dictionary directly, don't call it
            rem["status"] = stat

        # F
        if cn_num and redis_client.sismember("spe:cn_nums", int(cn_num)):
            reminder_list.append({
                "title": "nc:spe_cn_num",
                "content": f"{cn_num} REMOVED",
                "id": None,
                "is_all_business": 1,
                "is_global": 1,
                "market": ["USCN", "CACN"],
                "recurrent_interval": 0,
                "start_date": datetime(2000, 1, 1).isoformat(),
                "end_date": datetime(5000, 1, 1).isoformat(),
                "status": {
                    "is_processed": 0,
                    "last_processed_at": datetime.fromtimestamp(0)
                }
            })

        return reminder_list

    except Exception as e:
        self.update_state(state="FAILURE")
        raise Exception(f"Error processing reminder insertion: {str(e)}")

@celery.task(bind=True)
def process_op(self, op):
    db_manager = DatabaseManager()
    
    try:
        db_manager.insert_reminder_op(op)
        db_manager.insert_process_log(op)

        redis_client.hset(f"rem:{op['reminder_id']}:{op['acct_id']}",
            mapping={"is_processed": int(op["is_processed"]), "last_processed_at": op["processed_at"]}
        )
    except Exception as e:
        self.update_state(state="FAILURE")
        raise Exception(f"Error processing operation: {str(e)}")

@celery.task(bind=True)
def delete_reminder_task(self, reminder_id):
    db_manager = DatabaseManager()

    try:
        db_manager.delete_reminder(reminder_id)

        temp_pipe = redis_client.pipeline()
        temp_pipe.delete(f"rem:detail:{reminder_id}")
        for k in redis_client.scan_iter(f"rem:{reminder_id}:*"):
            temp_pipe.delete(k)
        temp_pipe.execute()

    except Exception as e:
        self.update_state(state="FAILURE")
        raise Exception(f"Error deleting reminderr: {str(e)}")

@celery.task(bind=True)
def create_acct_xlsx(self, reminder_id):
    try:
        temp_dir = tempfile.gettempdir()
        file_pattern = os.path.join(temp_dir, f"acct_{reminder_id}_*.xlsx")
        existing_files = glob.glob(file_pattern)

        if existing_files:
            latest_file = max(existing_files, key=os.path.getmtime)
            return latest_file

        try:
            with tempfile.NamedTemporaryFile(prefix=f'acct_{reminder_id}_', suffix='.xlsx', delete=False, dir=temp_dir) as tmp_file:
                file_path = tmp_file.name
        except Exception as e:
            self.update_state(state="FAILURE")
            raise Exception(f"Failed to create temporary file: {str(e)}")

        db_manager = DatabaseManager()
        acct_ids = db_manager.get_accts_by_reminder(reminder_id)

        wb = Workbook()
        ws = wb.active
        for i, acct_id in enumerate(acct_ids):
            ws.cell(row=i+1, column=1, value=acct_id)

        wb.save(file_path)
        delete_file.apply_async((file_path,), countdown=300)
        return file_path

    except Exception as e:
        # Clean up the temp file if created but failed before saving/scheduling deletion
        if 'file_path' in locals() and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                pass
        self.update_state(state="FAILURE", meta={'exc_type': type(e).__name__, 'exc_message': str(e)})
        raise Exception(f"Error in create_acct_xlsx for reminder {reminder_id}: {str(e)}")

@celery.task(bind=True)
def get_process_status(self, reminder_id):
    """
    Generates an Excel file containing the process log and account status for a given reminder.

    Args:
        reminder_id (int): The ID of the reminder.

    Returns:
        str: The path to the generated temporary Excel file.
             Returns False if an error occurs.
    """
    try:
        temp_dir = tempfile.gettempdir()
        file_pattern = os.path.join(temp_dir, f"status_{reminder_id}_*.xlsx")
        existing_files = glob.glob(file_pattern)

        if existing_files:
            latest_file = max(existing_files, key=os.path.getmtime)
            return latest_file

        db_manager = DatabaseManager()
        reminder_log_data = db_manager.get_reminder_log(reminder_id)
        acct_status_data = db_manager.get_all_acct_reminder_status(reminder_id) # Use the specific method

        if reminder_log_data is False or acct_status_data is False:
             self.update_state(state="FAILURE", meta={'exc_type': 'DatabaseError', 'exc_message': 'Failed to retrieve data from database.'})
             raise Exception(f"Failed to retrieve data for reminder_id {reminder_id}")

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "历史记录"
        if reminder_log_data:
            # Write header
            log_headers = list(reminder_log_data[0].keys())
            ws1.append(log_headers)
            # Write data rows
            for row_dict in reminder_log_data:
                # Ensure consistent order based on headers
                row_values = [row_dict.get(header) for header in log_headers]
                ws1.append(row_values)
        else:
            ws1.append(["无历史数据"]) # Placeholder if no data

        # Sheet 2: Account Status
        ws2 = wb.create_sheet(title="账户状态")
        if acct_status_data:
            # Write header
            status_headers = list(acct_status_data[0].keys())
            ws2.append(status_headers)
            # Write data rows
            for row_dict in acct_status_data:
                 # Ensure consistent order based on headers
                row_values = [row_dict.get(header) for header in status_headers]
                ws2.append(row_values)
        else:
            ws2.append(["无账户数据"])

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=f"status_{reminder_id}_") as tmp:
            file_path = tmp.name
            wb.save(file_path)

        delete_file.apply_async(args=[file_path], countdown=300)

        return file_path

    except Exception as e:
        self.update_state(state="FAILURE", meta={'exc_type': type(e).__name__, 'exc_message': str(e)})
        raise Exception(f"Error generating process status file for reminder {reminder_id}: {str(e)}")

@celery.task
def delete_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)

@celery.task(bind=True)
def search_reminders(self, search_type, **kwargs):
    """
    Consolidated task to search reminders by ID, creation date, or title.
    Converts business IDs to a comma-separated list of names.
    """
    db_manager = DatabaseManager()
    reminders = []

    try:
        if search_type == 'id':
            reminder_id = kwargs.get('reminder_id')
            if not reminder_id:
                raise ValueError("reminder_id is required for search_type 'id'")
            result = db_manager.get_reminder_by_id(reminder_id)
            if result:
                reminders = [result] # Wrap single result in a list for consistent processing
        elif search_type == 'date':
            start_date = kwargs.get('start_date')
            end_date = kwargs.get('end_date')
            if not end_date:
                end_date = datetime(5000, 1, 1)
            reminders = db_manager.get_reminders_by_create_date(start_date, end_date)
        elif search_type == 'title':
            title_query = kwargs.get('title_query')
            reminders = db_manager.get_reminders_by_title(title_query)
        else:
            raise ValueError(f"Invalid search_type: {search_type}")

        if not reminders:
            # Return None for ID search if not found, empty list otherwise
            return []

        processed_reminders = []
        for reminder in reminders:
            business_ids = reminder.pop('business_ids', []) # Remove business_ids key
            business_names = []
            if business_ids:
                for b_id in business_ids:
                    business_names.append(redis_client.get(f'business:id:{b_id}').decode('utf-8'))

            reminder['business_names'] = ','.join(sorted(list(set(business_names))))

            # Convert datetime objects to ISO strings for JSON serialization
            for key in ['start_date', 'expire_date', 'create_date', 'last_modified']:
                 if isinstance(reminder.get(key), datetime):
                     reminder[key] = reminder[key].isoformat()

            processed_reminders.append(reminder)

        return processed_reminders

    except ValueError as ve:
        self.update_state(state="FAILURE", meta={'exc_type': 'ValueError', 'exc_message': str(ve)})
        # Re-raise to ensure Celery marks task as failed
        raise Exception(f'Invalid input for search: {str(ve)}')
    except Exception as e:
        self.update_state(state="FAILURE", meta={'exc_type': type(e).__name__, 'exc_message': str(e)})
        # Log the full error for debugging
        print(f"Error in search_reminders task: {type(e).__name__} - {str(e)}")
        raise Exception(f'Error processing reminder search: {str(e)}')

#####################
#
# Reminder Endpoints
#
#####################
@app.route("/reminder/insert", methods=["POST"])
def insert_reminder():
    try:
        # Determine reminder data source based on content type
        if request.content_type.startswith('multipart/form-data'):
            reminder_json_str = request.form.get('reminder_data')
            reminder = json.loads(reminder_json_str)
        else:
            reminder = request.get_json()

        if not reminder:
            return jsonify({'message': 'No valid reminder data found'}), 400

        if not reminder.get('is_global', False):
            # Handle non-global reminder (file upload for acct_ids)
            if 'file' not in request.files:
                return jsonify({'message': 'No file part for non-global reminder'}), 400
            file = request.files['file']
            if file.filename == '':
                return jsonify({'message': 'No selected file'}), 400
            if file and file.filename.endswith('.xlsx'):
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=tempfile.gettempdir()) as tmp_file:
                    file.save(tmp_file.name)
                    file_path = tmp_file.name

                task = (get_acct_from_form.s(file_path) | insert_reminder_task.s(reminder)).apply_async()
                return jsonify({'task_id': task.id}), 202
            else:
                return jsonify({'message': 'Invalid file type. Please upload an XLSX file'}), 400
        else:
            task = insert_reminder_task.delay(None, reminder)
            return jsonify({'task_id': task.id}), 202

    except Exception as e:
        return jsonify({'message': f'Error processing request: {str(e)}'}), 500

@app.route("/reminder/acct", methods=['POST'])
def get_reminder():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'message': 'No data provided'}), 400

        required_fields = ['acct_id', 'market', 'business']
        if not all(data.get(field) for field in required_fields):
            return jsonify({'message': 'Missing required fields: acct_id, market, or business'}), 400

        task = get_reminders_for_acct.delay(data['acct_id'], data['market'], data['business'], data.get('cn_num'))
        return jsonify({'task_id': task.id}), 202

    except Exception as e:
        return jsonify({'message': f'Error processing request: {str(e)}'}), 500

@app.route("/reminder/op", methods=['POST'])
def operate_on_reminder():
    try:
        data = request.get_json()

        required_fields = ['acct_id', 'reminder_id', 'processed_by', 'processed_by_xf', 'is_processed']
        if not all(field in data for field in required_fields):
            return jsonify({'message': 'Missing required fields: acct_id, reminder_id, processed_by, or is_processed'}), 400

        if 'processed_at' not in data:
            data['processed_at'] = datetime.now().isoformat()

        task = process_op.delay(data)
        return jsonify({'task_id': task.id}), 202

    except Exception as e:
        return jsonify({'message': f'Error processing request: {str(e)}'}), 500

@app.route("/reminder/delete/<reminder_id>", methods=['DELETE'])
def delete_reminder(reminder_id):
    try:
        task = delete_reminder_task.delay(reminder_id)
        return jsonify({'task_id': task.id}), 202
    except Exception as e:
        return jsonify({'message': f'Error deleting reminder: {str(e)}'}), 500

@app.route("/reminder/search/id/<reminder_id>", methods=["GET"])
def get_reminders_by_id(reminder_id):
    """Starts a Celery task to search for a reminder by its ID."""
    try:
        task = search_reminders.delay(search_type='id', reminder_id=reminder_id)
        return jsonify({'task_id': task.id}), 202
    except Exception as e:
        return jsonify({'message': f'Error starting search task: {str(e)}'}), 500

@app.route("/reminder/search/createtime", methods=["GET"])
def get_reminders_by_createtime():
    """Starts a Celery task to search for reminders by creation date range."""
    try:
        # Read from query parameters
        start_date = request.args.get('start_date')
        if not start_date:
            return jsonify({'message': 'Missing start_date query parameters'}), 400
        task = search_reminders.delay(search_type='date', start_date=start_date, end_date=request.args.get('end_date'))
        return jsonify({'task_id': task.id}), 202
    except Exception as e:
        return jsonify({'message': f'Error starting search task: {str(e)}'}), 500

@app.route("/reminder/search/title", methods=["GET"])
def get_reminders_by_title():
    """Starts a Celery task to search for reminders by title."""
    try:
        # Read from query parameters
        title_query = request.args.get('title')
        if not title_query:
            return jsonify({'message': 'Missing title query parameter'}), 400
        if len(title_query) < 2:
            return jsonify({'message': 'title_query must be at least 2 characters long'}), 400

        task = search_reminders.delay(search_type='title', title_query=title_query)
        return jsonify({'task_id': task.id}), 202
    except Exception as e:
        return jsonify({'message': f'Error starting search task: {str(e)}'}), 500

@app.route("/reminder/getaccts/<reminder_id>", methods=['GET'])
def get_acct_list(reminder_id):
    try:
        task = create_acct_xlsx.delay(reminder_id)
        return jsonify({'task_id': task.id}), 202
    except Exception as e:
        return jsonify({'message': f'Error getting acct list: {str(e)}'}), 500

@app.route("/reminder/status/<reminder_id>", methods=['GET'])
def get_reminder_status_file(reminder_id):
    """Starts a Celery task to generate the status Excel file for a reminder."""
    try:
        # Validate reminder_id is an integer if needed
        try:
            int(reminder_id)
        except ValueError:
            return jsonify({'message': 'Invalid reminder ID format'}), 400

        task = get_process_status.delay(reminder_id)
        return jsonify({'task_id': task.id}), 202
    except Exception as e:
        return jsonify({'message': f'Error starting status file generation task: {str(e)}'}), 500

@app.route("/reminder/download/<task_id>", methods=['GET'])
def download_reminder_file_by_task(task_id):
    """
    Downloads the file generated by a specific Celery task (e.g., create_acct_xlsx),
    retrieving the file path from the task result.
    """
    task = celery.AsyncResult(task_id)

    if task.state == 'SUCCESS':
        file_path = task.result
        if not file_path or not isinstance(file_path, str):
             app.logger.error(f"Task {task_id} succeeded but result is not a valid file path: {file_path}")
             return jsonify({'message': 'Task result invalid'}), 500

        # Security Check: Ensure the path is within the system's temp directory and is an xlsx file
        requested_path_abs = os.path.abspath(file_path)

        if not requested_path_abs.endswith('.xlsx'):
            app.logger.warning(f"Task {task_id} result path potentially unsafe: {file_path}")
            return jsonify({'message': 'Invalid or unsafe file path in task result'}), 400

        if not os.path.exists(requested_path_abs):
            app.logger.error(f"File from task {task_id} not found at path: {requested_path_abs}")
            return jsonify({'message': 'File not found or expired'}), 404

        try:
            # Use the original filename for the download
            download_name = os.path.basename(requested_path_abs)

            return send_file(requested_path_abs, as_attachment=True, download_name=download_name)
        except Exception as e:
            app.logger.error(f"Error sending file {requested_path_abs} for task {task_id}: {str(e)}")
            return jsonify({'message': f'Error sending file: {str(e)}'}), 500

    elif task.state == 'PENDING' or task.state == 'STARTED' or task.state == 'RETRY':
        return jsonify({'message': 'File generation is still in progress'}), 202 # Accepted, but not ready
    else: # FAILURE or other terminal state
        app.logger.error(f"Task {task_id} failed or is in unexpected state {task.state}. Info: {task.info}")
        return jsonify({'message': f'File generation failed or task error: {task.state}'}), 500

@app.route('/reminder/ReminderManagement.html')
def rem_manage():
    return send_file('ReminderManagement.html')
#####################
#
# Plan Endpoints
#
#####################
@app.route("/plan/update", methods=['POST'])
def insert_plan():
    plan = request.get_json()
    if not plan:
        return jsonify({'message': 'Invalid plan data'}), 400
    task = sink_plan_to_db.delay([plan]) # defaulting to update to overwrite
    return jsonify({'task_id': task.id}), 202

@app.route('/plan/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400
    if file and file.filename.endswith('.xlsx'):
        try:
            # Use tempfile for secure temporary file creation
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=tempfile.gettempdir()) as tmp_file:
                file.save(tmp_file.name)
                file_path = tmp_file.name

            task = (process_excel_data.s(file_path) | sink_plan_to_db.s()).apply_async()

            return jsonify({'task_id': task.id}), 202
        except Exception as e:
            return jsonify({'message': f'Error processing file: {str(e)}'}), 500
    else:
        return jsonify({'message': 'Invalid file type. Please upload an XLSX file'}), 400

@app.route('/plan/<plan_id>', methods=['GET'])
def get_plan(plan_id):
    try:
        task = get_plan_info.delay(plan_id) 
        return jsonify({"task_id": task.id}), 202
    
    except Exception as e:
        return jsonify({'message': f'Error loading plan: {plan_id}'}), 500

# Plan update frontend
@app.route('/plan/PlanManagement.html')
def plan_manage():
    return send_file('PlanManagement.html')

#####################
#
# Other Endpoints
#
#####################   
@app.route('/tasks/<task_id>', methods=['GET'])
def get_task_status(task_id):
    task = celery.AsyncResult(task_id)
    if task.state == 'PENDING':
        response = {
            'status': task.state
        }
    elif task.state == 'SUCCESS':
        response = {
            'status': task.state,
            'result': task.result
        }
    else:
        # something went wrong in the background job
        response = {
            'status': task.state,
            'message': str(task.info),  # this is the exception raised
        }
    return jsonify(response), 200